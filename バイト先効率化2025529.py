import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime, timedelta

# グローバル変数 name の初期化
def excel_serial_to_date(serial_number):
    base_date = datetime(1899, 12, 30)  # Excelは1900年1月1日を「1」と数える
    return base_date + timedelta(days=serial_number)
name = ""
last_list = []

def submit_name():
    global name
    # 入力された名前を取得してグローバル変数に代入
    name = entry.get()
    print(f"提出された名前: {name}")
    # 入力欄とボタンを削除
    entry.destroy()
    submit_button.destroy()
    open_window()
    

# ウィンドウの作成
root = tk.Tk()
root.title("名前の提出")
root.geometry("300x100")

# 名前入力欄
entry = tk.Entry(root)
entry.pack(pady=10)

# 提出ボタン
submit_button = tk.Button(root, text="提出", command=submit_name)
submit_button.pack()

file_path = ""

def select_file():
    global file_path
    # Excelファイルだけを対象にファイル選択ダイアログを表示
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")],
        title="Excelファイルを選択"
    )
    if file_path:
        print(f"選択されたファイル: {file_path}")

def submit_file():
    if file_path:
        print(f"提出されたファイルパス: {file_path}")
        select_button.destroy()
        submit_button.destroy()
        file_parse()

def open_window():
    # ファイル選択ボタン
    global select_button, submit_button
    select_button = tk.Button(root, text="ファイルを選択", command=select_file)
    select_button.pack(pady=10)

    # 提出ボタン
    submit_button = tk.Button(root, text="提出", command=submit_file)
    submit_button.pack(pady=10)


def file_parse():
    global file_path
    try:
        wb = load_workbook(file_path)
        ws = wb.active  # 最初のシートをアクティブに

        rows = 0
        row_num = 6
        while True:
            cell_value = ws[f"B{row_num}"].value
            if cell_value is None:
                break
            rows += 1
            row_num += 1

        data_list = []
        row_number = 6
        for _ in range(rows):
            columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            data_dict = {}
            for col in columns:
                cell_value = ws[f"{col}{row_number}"].value
                data_dict[col] = cell_value
            data_dict["row_number"] = row_number  # 行番号を追加
            data_list.append(data_dict)
            row_number += 1

        unique_data_list = []
        for data in data_list:
            if data['H'] == '店舗間移動\u3000\u3000\u3000届け先をG列に記載→':
                unique_data_list.append(data)
        global last_list
       
        for data in unique_data_list:
            if data["K"] is None:
                last_list.append(data)

        print("最終的なリスト:", last_list[0])

    except Exception as e:
        print("エラーが発生しました:", e)
        
    loopaction()
counter=0
def loopaction():
    
    global counter, input_number, date, sender, receiver, how_many,button,last_list,root
    real_date=excel_serial_to_date(int(last_list[counter]["D"]))
    date = tk.Label(root, text=f"移動日: {real_date.strftime("%Y-%m-%d")}")
    date.pack()
    sender = tk.Label(root, text=f"送り元: {last_list[counter]["B"]}")
    sender.pack()
    receiver = tk.Label(root, text=f"送り先: {last_list[counter]["I"]}")
    receiver.pack()
    how_many = tk.Label(root, text=f"送った個数: {last_list[counter]["G"]}")
    how_many.pack()
    global input_number, button
    input_number = tk.Entry(root)
    input_number.pack(pady=10)
    button = tk.Button(root, text="次へ", command=next_action)
    button.pack(pady=10)

def next_action():
    global counter, date, sender, receiver, how_many, input_number,button, last_list

    #ここにエクセル入力する関数を追加する
    excel_input()

    date.destroy()
    sender.destroy()        
    receiver.destroy()
    how_many.destroy()
    input_number.destroy()
    button.destroy()
    counter += 1
    if counter < len(last_list):
        
        loopaction()
    
    else:
        label= tk.Label(root, text="全てのデータを処理しました。")
        label.pack(pady=10)

def excel_input():
    global input_number, last_list, counter,name
    #last_list[counter]についての処理を記述。ループは関数ごと行う。
    try:
        wb = load_workbook(file_path)
        ws = wb.active  # 最初のシートをアクティブに
        # 入力値を取得
        input_value = input_number.get()
        
        # 行番号を取得
        row_number = last_list[counter]["row_number"]
        
        # K列に入力値を設定
        ws[f"K{row_number}"] = input_value

        ws[f"J{row_number}"] = name  # 名前をJ列に入力
        
        # ワークブックを保存
        wb.save(file_path)
        
        print(f"行 {row_number} の 処理番号に '{input_value}' を入力しました。")
        
    except Exception as e:
        print("エラーが発生しました:", e)
    


        
    






# メインループ開始
root.mainloop()
